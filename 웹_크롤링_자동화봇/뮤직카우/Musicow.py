import requests
from bs4 import BeautifulSoup
import re
import json
import openpyxl
import datetime

# Workbook 생성
wb = openpyxl.Workbook()
# Sheet 생성
wb.create_sheet('1. 기본정보', 0)
wb.create_sheet('2. 저작권료', 1)
wb.create_sheet('3. 상세저작권료', 2)
# Sheet 활성
sheet = wb['1. 기본정보']
sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 11
sheet.column_dimensions['D'].width = 11
sheet.column_dimensions['E'].width = 20
sheet.column_dimensions['K'].width = 18
sheet.column_dimensions['L'].width = 40

sheet.append(["곡명", "가수", "인접권여부", "공표일자", "2차적저작물작성권", "현재가", "저작권료", "수익률", "총주수", "유통주수", "시가총액", "기타 주요사항"])

sheet = wb['2. 저작권료']
sheet.append(["곡명", "가수"
                 , "2021-12", "2021-11", "2021-10", "2021-09", "2021-08", "2021-07", "2021-06", "2021-05", "2021-04",
              "2021-03", "2021-02", "2021-01"
                 , "2020-12", "2020-11", "2020-10", "2020-09", "2020-08", "2020-07", "2020-06", "2020-05", "2020-04",
              "2020-03", "2020-02", "2020-01"
                 , "2019-12", "2019-11", "2019-10", "2019-09", "2019-08", "2019-07", "2019-06", "2019-05", "2019-04",
              "2019-03", "2019-02", "2019-01"
                 , "2018-12", "2018-11", "2018-10", "2018-09", "2018-08", "2018-07", "2018-06", "2018-05", "2018-04",
              "2018-03", "2018-02", "2018-01"
                 , "2017-12", "2017-11", "2017-10", "2017-09", "2017-08", "2017-07", "2017-06", "2017-05", "2017-04",
              "2017-03", "2017-02", "2017-01"])

sheet = wb['3. 상세저작권료']
sheet.append(["곡명", "가수", "인접권여부", "1년저작권료", "방송", "전송", "복제", "공연", "해외", "기타", "디음송"])

years = ["2021", "2020", "2019", "2018", "2017"]
months = ["12", "11", "10", "09", "08", "07", "06", "05", "04", "03", "02", "01"]

base_url = 'https://www.musicow.com/song/{}?tab=info'
headers = {
    'User-Agent': 'Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57'}

Max_ID = input("숫자를 입력하세요: ")

for n in range(int(Max_ID)):
    # 하나씩 추출
    url = base_url.format(n + 1)
    res = requests.get(url, headers=headers).text
    soup = BeautifulSoup(res, "lxml")
    pattern = re.compile(r"arr_amt_royalty_ym\['graph1'\] = (.*?);$", re.MULTILINE | re.DOTALL)
    script = soup.find("script", text=pattern)
    dic = pattern.search(script.string).group(1)
    string_to_dict = json.loads(dic)
    st_len = len(string_to_dict)
    # 곡에 해당하는 site가 있을 경우,
    if st_len != 0:
        # 제목 및 아티스트
        곡명 = soup.select("#page_market > div.container > div.song_header > div.information > p > strong")[0].text.strip()
        가수 = soup.select("#page_market > div.container > div.song_header > div.information > em")[0].text.strip()

        # 1. 기본정보
        sheet = wb['1. 기본정보']
        try:
            soup.select("#page_market > div.container > div.song_header > div.information > p > i")[0].text
            인접권여부 = 'O'
        except:
            인접권여부 = 'X'

        공표일자 = soup.select("div.lst_copy_info > dl > dd")[0].text.strip()
        if (인접권여부 == 'X'):
            이차적저작물작성권 = soup.select("div.lst_copy_info > dl > dd > p")[1].text.split("(")[1].replace(")", "").strip()
        else:
            이차적저작물작성권 = ""

        현재가 = int(soup.select(
            "#page_market > div.container > div.song_header > div.price_wrap > dl:nth-child(1) > dd > strong")[
                      0].text.split(" ")[0].replace(",", ""))

        저작권료 = \
        soup.select("#page_market > div.container > div.song_header > div.price_wrap > dl:nth-child(3) > dd > span")[
            0].text.split(" ")[0]
        if ("," in list(저작권료)):
            저작권료 = int(저작권료.replace(",", ""))
        else:
            저작권료 = int(저작권료)

        수익률 = \
        soup.select("#page_market > div.container > div.song_header > div.price_wrap > dl:nth-child(3) > dd > span")[
            0].text.split(" ")[1].replace("(", "").replace(")", "")

        총주수 = int(soup.select("div.lst_copy_info > dl > dd > p")[0].text.split("/")[1].replace(",", ""))
        길이 = len(soup.select(".lst_numb_card dl dd"))
        if (길이 < 11):
            유통주수 = soup.select(".lst_numb_card dl dd")[3].text.split(" ")[0]
            if ("," in list(유통주수)):
                유통주수 = int(유통주수.replace(",", ""))
            else:
                유통주수 = int(유통주수)
        else:
            유통주수_1차 = soup.select(".lst_numb_card dl dd")[3].text.split(" ")[0]
            유통주수_2차 = soup.select(".lst_numb_card dl dd")[7].text.split(" ")[0]
            if ("," in list(유통주수_1차)):
                유통주수_1차 = int(유통주수_1차.replace(",", ""))
            else:
                유통주수_1차 = int(유통주수_1차)
            if ("," in list(유통주수_2차)):
                유통주수_2차 = int(유통주수_2차.replace(",", ""))
            else:
                유통주수_2차 = int(유통주수_2차)

            유통주수 = 유통주수_1차 + 유통주수_2차

        시가총액 = 현재가 * 총주수
        기타_주요사항 = soup.select("div.lst_copy_info > dl > dd > ul")[0].text.strip()

        sheet.append([곡명, 가수, 인접권여부, 공표일자, 이차적저작물작성권, 현재가, 저작권료, 수익률, 총주수, 유통주수, 시가총액, 기타_주요사항])

        # 2. 저작권료
        sheet = wb['2. 저작권료']
        # 배열 초기화
        temp_list = [];
        for year in years:
            for month in months:
                try:
                    value = string_to_dict[year][month]
                    temp_list.append(value)
                except:
                    temp_list.append('-1')

        sheet.append(
            [곡명, 가수, temp_list[0], temp_list[1], temp_list[2], temp_list[3], temp_list[4], temp_list[5], temp_list[6],
             temp_list[7], temp_list[8], temp_list[9],
             temp_list[10], temp_list[11], temp_list[12], temp_list[13], temp_list[14], temp_list[15], temp_list[16],
             temp_list[17], temp_list[18], temp_list[19],
             temp_list[20], temp_list[21], temp_list[22], temp_list[23], temp_list[24], temp_list[25], temp_list[26],
             temp_list[27], temp_list[28], temp_list[29],
             temp_list[30], temp_list[31], temp_list[32], temp_list[33], temp_list[34], temp_list[35], temp_list[36],
             temp_list[37], temp_list[38], temp_list[39],
             temp_list[40], temp_list[41], temp_list[42], temp_list[43], temp_list[44], temp_list[45], temp_list[46],
             temp_list[47], temp_list[48], temp_list[49],
             temp_list[50], temp_list[51], temp_list[52], temp_list[53], temp_list[54], temp_list[55], temp_list[56],
             temp_list[57], temp_list[58], temp_list[59]])

        # 3. 상세저작권료
        sheet = wb['3. 상세저작권료']

        방송 = int(soup.select(
            "#song_info_royalty > div.card_body > div > div > div:nth-child(2) > div > div.old_money > div.tbl_flex > dl:nth-child(1) > dd")[
            0].text.split("원")[0].replace(",", ""))
        전송 = int(soup.select(
            "#song_info_royalty > div.card_body > div > div > div:nth-child(2) > div > div.old_money > div.tbl_flex > dl:nth-child(2) > dd")[
            0].text.split("원")[0].replace(",", ""))
        복제 = int(soup.select(
            "#song_info_royalty > div.card_body > div > div > div:nth-child(2) > div > div.old_money > div.tbl_flex > dl:nth-child(3) > dd")[
            0].text.split("원")[0].replace(",", ""))
        공연 = int(soup.select(
            "#song_info_royalty > div.card_body > div > div > div:nth-child(2) > div > div.old_money > div.tbl_flex > dl:nth-child(4) > dd")[
            0].text.split("원")[0].replace(",", ""))
        해외 = int(soup.select(
            "#song_info_royalty > div.card_body > div > div > div:nth-child(2) > div > div.old_money > div.tbl_flex > dl:nth-child(5) > dd")[
            0].text.split("원")[0].replace(",", ""))
        기타 = int(soup.select(
            "#song_info_royalty > div.card_body > div > div > div:nth-child(2) > div > div.old_money > div.tbl_flex > dl:nth-child(6) > dd")[
            0].text.split("원")[0].replace(",", ""))
        if(인접권여부 == "O"):
            디음송 = soup.select(
                "#song_info_royalty > div.card_body > div > div > div:nth-child(2) > div > div.old_money > div.tbl_flex > dl:nth-child(7) > dd")[
            0].text.split("원")[0].replace(",", "")
        else:
            디음송 = 0
        일년저작권료 = 방송 + 전송 + 복제 + 공연 + 해외 + 기타 + 디음송

        sheet.append([곡명, 가수, 인접권여부, 일년저작권료, 방송, 전송, 복제, 공연, 해외, 기타, 디음송])
        print("good {0}".format(n))

dt_now = datetime.datetime.now()
formattedDate = dt_now.strftime("%Y%m%d_%H%M")
Filename = "Musicow_" + formattedDate + ".xlsx"

wb.save(Filename)