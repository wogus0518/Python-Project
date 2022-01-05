import random
import openpyxl

wb = openpyxl.load_workbook('program_data_ex.xlsx')

# 시트 얻기 ==> 'list_all'이름 시트를 listAll로 지정
listAll = wb['list_all']


#######################################################################################################################################
# sheet 안에서 데이터가 있는 범위만 지정
listAll = listAll['A2':'K285']
# 참가자 수 설정
num_submit = 284
# Type은 몇 번 열인가?
type_num = 6
########################################################################################################################################


# 가로 한 줄에 데이터 몇개 들어있나
data_len = len(listAll[0])
                                                                                                                                                                                                      ##########Type변수 엑셀에서 몇번째 행인지 수정 필요##########
#라일락, 데이지, clover 리스트 생성
lilac_list, daisy_list, clover_list = [], [], []

# listAll => lilac_list
for people_idx in range(num_submit):
    if listAll[people_idx][type_num].value == 'lilac':
        lilac_list.append(listAll[people_idx])
print('라일락 신청자 분류 완료!!!')


# listAll => daisy_list
for people_idx in range(num_submit):
    if listAll[people_idx][type_num].value == 'daisy':
        daisy_list.append(listAll[people_idx])
print('데이지 신청자 분류 완료!!!')


# listAll => clover_list
for people_idx in range(num_submit):
    if listAll[people_idx][type_num].value == 'clover':
        clover_list.append(listAll[people_idx])
print('클로버 신청자 분류 완료!!!')

##########################################################################################################################################
##########################################################################################################################################


keys = ('Nic', 'Name', 'Phone', 'User_univ', 'User_age', 'Sex', \
        'Type', 'Dating_univ', 'Dating_age', 'agreement', 'Ticket')

# lilac으로 분류한 사람들 dictionary화
num_lilac = len(lilac_list)
lilac_dict_list = []
for idx in range(num_lilac):
    people_info_lilac = lilac_list[idx]
    lilac_dict = {key: person_info_lilac for key, person_info_lilac in zip(keys, people_info_lilac)}
    lilac_dict_list.append(lilac_dict)
# daisy으로 분류한 사람들 dictionary화
num_daisy = len(daisy_list)
daisy_dict_list = []
for idx in range(num_daisy):
    people_info_daisy = daisy_list[idx]
    daisy_dict = {key: person_info_daisy for key, person_info_daisy in zip(keys, people_info_daisy)}
    daisy_dict_list.append(daisy_dict)
# clover으로 분류한 사람들 dictionary화
num_clover = len(clover_list)
clover_dict_list = []
for idx in range(num_clover):
    people_info_clover = clover_list[idx]
    clover_dict = {key: person_info_clover for key, person_info_clover in zip(keys, people_info_clover)}
    clover_dict_list.append(clover_dict)


##########################################################################################################################################
##########################################################################################################################################


# 라일락에서 돈 낸 남자들
lilac_matching_boy = list()
for idx in range(num_lilac):
    lilac_person = lilac_dict_list[idx]
    if lilac_person['Sex'].value == 'boy' and lilac_person['Ticket'].value != 0:
        lilac_matching_boy.append(lilac_person)
# 라일락에서 돈 낸 여자들
lilac_matching_girl = list()
for idx in range(num_lilac):
    lilac_person = lilac_dict_list[idx]
    if lilac_person['Sex'].value == 'girl' and lilac_person['Ticket'].value != 0:
        lilac_matching_girl.append(lilac_person)
# 라일락 돈 낸 남자+여자
lilac_matching = lilac_matching_boy + lilac_matching_girl
print('라일락 번호 뽑기 신청자 분류 완료!!!')

# 데이지에서 돈 낸 남자들
daisy_matching_boy = list()
for idx in range(num_daisy):
    daisy_person = daisy_dict_list[idx]
    if daisy_person['Sex'].value == 'boy' and daisy_person['Ticket'].value != 0:
        daisy_matching_boy.append(daisy_person)
# 데이지에서 돈 낸 여자들
daisy_matching_girl = list()
for idx in range(num_daisy):
    daisy_person = daisy_dict_list[idx]
    if daisy_person['Sex'].value == 'girl' and daisy_person['Ticket'].value != 0:
        daisy_matching_girl.append(daisy_person)
# 데이지 돈 낸 남자+여자
daisy_matching = daisy_matching_boy + daisy_matching_girl
print('데이지 번호 뽑기 신청자 분류 완료!!!')

# clover에서 돈 낸 남자들
clover_matching_boy = list()
for idx in range(num_clover):
    clover_person = clover_dict_list[idx]
    if clover_person['Sex'].value == 'boy' and clover_person['Ticket'].value != 0:
        clover_matching_boy.append(clover_person)
# clover에서 돈 낸 여자들
clover_matching_girl = list()
for idx in range(num_clover):
    clover_person = clover_dict_list[idx]
    if clover_person['Sex'].value == 'girl' and clover_person['Ticket'].value != 0:
        clover_matching_girl.append(clover_person)
# clover 돈 낸 남자+여자
clover_matching = clover_matching_boy + clover_matching_girl
print('클로버 번호 뽑기 신청자 분류 완료!!!')


##########################################################################################################################################
##########################################################################################################################################


# 엑셀 파일에 저장하기
## 결과값을 저장할 새로운 엑셀 파일 만들기
### naming하고 시트 추가하기
wb = openpyxl.Workbook()

listLilac = wb.active
listLilac.title = "list_lilac"
listDaisy = wb.create_sheet('list_daisy')
listclover = wb.create_sheet('list_clover')
lilacMoney = wb.create_sheet('lilac_money')
daisyMoney = wb.create_sheet('daisy_money')
cloverMoney = wb.create_sheet('clover_money')
lilacResult = wb.create_sheet('lilac_result')
daisyResult = wb.create_sheet('daisy_result')
cloverResult = wb.create_sheet('clover_result')


# 분류한 명단을 각각의 시트에 저장하는 함수 코드
## list와 dictionary로 다르게 저장되어 있는 명단을 각각 따로 만들어야함

sorted_lists = [lilac_list, daisy_list, clover_list, lilac_matching, daisy_matching, clover_matching]
worksheets = [listLilac, listDaisy, listclover, lilacMoney, daisyMoney, cloverMoney]

def save_func_list(sorted_list, worksheet):
    for colum_idx in range(len(sorted_list)):
        for row_idx in range(data_len):
            worksheet.cell(colum_idx + 1, row_idx + 1).value = sorted_list[colum_idx][row_idx].value

def save_func_dict(sorted_list, worksheet):
    for colum_idx in range(len(sorted_list)):
        for row_idx in range(data_len):
            worksheet.cell(colum_idx + 1, row_idx + 1).value = sorted_list[colum_idx][keys[row_idx]].value


save_func_list(lilac_list, listLilac)
save_func_list(daisy_list, listDaisy)
save_func_list(clover_list, listclover)
save_func_dict(lilac_matching, lilacMoney)
save_func_dict(daisy_matching, daisyMoney)
save_func_dict(clover_matching, cloverMoney)


wb.save('result40000.xlsx')
print('신청자 분류 후 엑셀 시트 작성 완료!!!')
print(f"=======================신청자 수======================\n"
      f"          |     남자     |     여자     \n"
      f"   라일락  |    {len(lilac_matching_boy)}명      |     {len(lilac_matching_girl)}명\n"
      f"   데이지  |    {len(daisy_matching_boy)}명      |     {len(daisy_matching_girl)}명\n"
      f"   클로버  |    {len(clover_matching_boy)}명      |     {len(clover_matching_girl)}명\n")

##########################################################################################################################################
##########################################################################################################################################


# 매칭 시켜줘야하는 사람의 조건에 맞는 사람들만 선별해서 하나의 리스트에 넣는 코드
keys = ('Nic', 'Name', 'Phone', 'User_univ', 'User_age', 'Sex', \
        'Type', 'Dating_univ', 'Dating_age', 'agreement', 'Ticket')

print('라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작')
print('라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작')
print('라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작')

for i in range(len(lilac_matching)):

    paid_person = lilac_matching[i]
    sex_condition = paid_person['Sex'].value
    univ_condition = paid_person['Dating_univ'].value
    age_condition = paid_person['Dating_age'].value

    pass_sex = list()
    pass_sex_univ = list()
    pass_sex_univ_age = list()

    print('\n')
    print(f"{paid_person['Name'].value} 매칭 시작 / 티켓 수: {paid_person['Ticket'].value}")

    # 성별 Pass
    for idx in range(len(lilac_dict_list)):
        candidate = lilac_dict_list[idx]
        if candidate['Sex'].value != sex_condition:
            pass_sex.append(candidate)
    '''
    print('1그룹')
    for row in pass_sex:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")
    '''
    if not pass_sex:
        print(f"{paid_person['Name'].value}님 ====> 1그룹 없음!!!")

    ## 성별 && 학교 Pass
    for idx in range(len(pass_sex)):
        candidate = pass_sex[idx]
        if univ_condition == 'myUniv' and paid_person['User_univ'].value == candidate['User_univ'].value:
            pass_sex_univ.append(candidate)

        elif univ_condition == 'dnt_m':
            if candidate['Dating_univ'].value == 'dnt_m':
                pass_sex_univ.append(candidate)
            elif candidate['Dating_univ'].value == 'myUniv':
                if paid_person['User_univ'].value == candidate['User_univ'].value:
                    pass_sex_univ.append(candidate)
            elif candidate['Dating_univ'].value == 'Oth_Univ':
                if candidate['User_univ'].value != paid_person['User_univ'].value:
                    pass_sex_univ.append(candidate)

        elif univ_condition == 'Oth_Univ':
            if candidate['Dating_univ'].value == 'Oth_Univ':
                if candidate['User_univ'].value != paid_person['User_univ'].value:
                    pass_sex_univ.append(candidate)
            elif candidate['Dating_univ'].value == 'dnt_m':
                if candidate['User_univ'].value != paid_person['User_univ'].value:
                    pass_sex_univ.append(candidate)

    # print('2그룹')
    # for row in pass_sex_univ:
    #     print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")

    if not pass_sex_univ:
        print(f"{paid_person['Name'].value}님 ====> 2그룹 없음!!!")

    ### 성별 && 학교 && 나이 Pass
    for idx in range(len(pass_sex_univ)):
        candidate = pass_sex_univ[idx]

        if age_condition == 'dnt_m':
            if candidate['Dating_age'].value == 'dnt_m':
                pass_sex_univ_age.append(candidate)
            elif candidate['Dating_age'].value != 'dnt_m':
                if candidate['Dating_age'].value == paid_person['User_age'].value:
                    pass_sex_univ_age.append(candidate)
        elif age_condition != 'dnt_m':
            if candidate['Dating_age'].value == 'dnt_m':
                if age_condition == candidate['User_age'].value:
                    pass_sex_univ_age.append(candidate)
            elif candidate['Dating_age'].value != 'dnt_m':
                if age_condition == candidate['User_age'].value:
                    if candidate['Dating_age'].value == paid_person['User_age'].value:
                        pass_sex_univ_age.append(candidate)



    print('3그룹')
    for row in pass_sex_univ_age:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")

    if not pass_sex_univ_age:
        print(f"{paid_person['Name'].value}님 ====> 3그룹 없음!!!")

    # ticket 수 만큼 후보자 항목에서 랜덤 추출(중복 안되게 할 것)
    ticket = paid_person['Ticket'].value
    result = random.sample(pass_sex_univ_age, ticket)
    print(result[0]['Phone'].value)
    for i


    # 매칭 결과 console print
    print(f"{paid_person['Name'].value} 매칭 결과")
    for row in result:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")































