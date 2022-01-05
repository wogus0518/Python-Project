import random
import openpyxl
import sys

sys.stdout = open('stdout.txt', 'w')
wb = openpyxl.load_workbook('program_data_ex.xlsx')

# 시트 얻기 ==> 'list_all'이름 시트를 listAll로 지정
listAll = wb['list_all']



#######################################################################################################################################
# sheet 안에서 데이터가 있는 범위만 지정
listAll = listAll['A2':'K285']
# 참가자 수 설정
num_submit = 284
# Type은 몇 번 열인가?
type_num = 6
########################################################################################################################################


# 가로 한 줄에 데이터 몇개 들어있나
data_len = len(listAll[0])
                                                                                                                                                                                                      ##########Type변수 엑셀에서 몇번째 행인지 수정 필요##########
#라일락, 데이지, clover 리스트 생성
lilac_list, daisy_list, clover_list = [], [], []

# listAll => lilac_list
for people_idx in range(num_submit):
    if listAll[people_idx][type_num].value == 'lilac':
        lilac_list.append(listAll[people_idx])
print('라일락 신청자 분류 완료!!!')

# listAll => daisy_list
for people_idx in range(num_submit):
    if listAll[people_idx][type_num].value == 'daisy':
        daisy_list.append(listAll[people_idx])
print('데이지 신청자 분류 완료!!!')

# listAll => clover_list
for people_idx in range(num_submit):
    if listAll[people_idx][type_num].value == 'clover':
        clover_list.append(listAll[people_idx])
print('클로버 신청자 분류 완료!!!')


##########################################################################################################################################
##########################################################################################################################################


keys = ('Nic', 'Name', 'Phone', 'User_univ', 'User_age', 'Sex', \
        'Type', 'Dating_univ', 'Dating_age', 'agreement', 'Ticket')

# lilac으로 분류한 사람들 dictionary화
num_lilac = len(lilac_list)
lilac_dict_list = []
for idx in range(num_lilac):
    people_info_lilac = lilac_list[idx]
    lilac_dict = {key: person_info_lilac for key, person_info_lilac in zip(keys, people_info_lilac)}
    lilac_dict_list.append(lilac_dict)
# daisy으로 분류한 사람들 dictionary화
num_daisy = len(daisy_list)
daisy_dict_list = []
for idx in range(num_daisy):
    people_info_daisy = daisy_list[idx]
    daisy_dict = {key: person_info_daisy for key, person_info_daisy in zip(keys, people_info_daisy)}
    daisy_dict_list.append(daisy_dict)
# clover으로 분류한 사람들 dictionary화
num_clover = len(clover_list)
clover_dict_list = []
for idx in range(num_clover):
    people_info_clover = clover_list[idx]
    clover_dict = {key: person_info_clover for key, person_info_clover in zip(keys, people_info_clover)}
    clover_dict_list.append(clover_dict)


##########################################################################################################################################
##########################################################################################################################################


# 라일락에서 돈 낸 남자들
lilac_matching_boy = list()
for idx in range(num_lilac):
    lilac_person = lilac_dict_list[idx]
    if lilac_person['Sex'].value == 'boy' and lilac_person['Ticket'].value != 0:
        lilac_matching_boy.append(lilac_person)
# 라일락에서 돈 낸 여자들
lilac_matching_girl = list()
for idx in range(num_lilac):
    lilac_person = lilac_dict_list[idx]
    if lilac_person['Sex'].value == 'girl' and lilac_person['Ticket'].value != 0:
        lilac_matching_girl.append(lilac_person)
# 라일락 돈 낸 남자+여자
lilac_matching = lilac_matching_boy + lilac_matching_girl
print('라일락 번호 뽑기 신청자 분류 완료!!!')

# 데이지에서 돈 낸 남자들
daisy_matching_boy = list()
for idx in range(num_daisy):
    daisy_person = daisy_dict_list[idx]
    if daisy_person['Sex'].value == 'boy' and daisy_person['Ticket'].value != 0:
        daisy_matching_boy.append(daisy_person)
# 데이지에서 돈 낸 여자들
daisy_matching_girl = list()
for idx in range(num_daisy):
    daisy_person = daisy_dict_list[idx]
    if daisy_person['Sex'].value == 'girl' and daisy_person['Ticket'].value != 0:
        daisy_matching_girl.append(daisy_person)
# 데이지 돈 낸 남자+여자
daisy_matching = daisy_matching_boy + daisy_matching_girl
print('데이지(이성) 번호 뽑기 신청자 분류 완료!!!')

# clover에서 돈 낸 남자들
clover_matching_boy = list()
for idx in range(num_clover):
    clover_person = clover_dict_list[idx]
    if clover_person['Sex'].value == 'boy' and clover_person['Ticket'].value != 0:
        clover_matching_boy.append(clover_person)
# clover에서 돈 낸 여자들
clover_matching_girl = list()
for idx in range(num_clover):
    clover_person = clover_dict_list[idx]
    if clover_person['Sex'].value == 'girl' and clover_person['Ticket'].value != 0:
        clover_matching_girl.append(clover_person)
# clover 돈 낸 남자+여자
clover_matching = clover_matching_boy + clover_matching_girl
print('데이지(동성) 번호 뽑기 신청자 분류 완료!!!')


##########################################################################################################################################
##########################################################################################################################################


# 엑셀 파일에 저장하기
## 결과값을 저장할 새로운 엑셀 파일 만들기
### naming하고 시트 추가하기
wb = openpyxl.Workbook()

listLilac = wb.active
listLilac.title = "list_lilac"
listDaisy = wb.create_sheet('list_daisy')
listclover = wb.create_sheet('list_clover')
lilacMoney = wb.create_sheet('lilac_money')
daisyMoney = wb.create_sheet('daisy_money')
cloverMoney = wb.create_sheet('clover_money')
lilacResult = wb.create_sheet('lilac_result')
daisyResult = wb.create_sheet('daisy_result')
cloverResult = wb.create_sheet('clover_result')


# 분류한 명단을 각각의 시트에 저장하는 함수 코드
## list와 dictionary로 다르게 저장되어 있는 명단을 각각 따로 만들어야함

sorted_lists = [lilac_list, daisy_list, clover_list, lilac_matching, daisy_matching, clover_matching]
worksheets = [listLilac, listDaisy, listclover, lilacMoney, daisyMoney, cloverMoney]

def save_func_list(sorted_list, worksheet):
    for colum_idx in range(len(sorted_list)):
        for row_idx in range(data_len):
            worksheet.cell(colum_idx + 1, row_idx + 1).value = sorted_list[colum_idx][row_idx].value

def save_func_dict(sorted_list, worksheet):
    for colum_idx in range(len(sorted_list)):
        for row_idx in range(data_len):
            worksheet.cell(colum_idx + 1, row_idx + 1).value = sorted_list[colum_idx][keys[row_idx]].value


save_func_list(lilac_list, listLilac)
save_func_list(daisy_list, listDaisy)
save_func_list(clover_list, listclover)
save_func_dict(lilac_matching, lilacMoney)
save_func_dict(daisy_matching, daisyMoney)
save_func_dict(clover_matching, cloverMoney)


wb.save('result40000.xlsx')
print('신청자 분류 후 엑셀 시트 작성 완료!!!')


##########################################################################################################################################
##########################################################################################################################################


# 매칭 시켜줘야하는 사람의 조건에 맞는 사람들만 선별해서 하나의 리스트에 넣는 코드
keys = ('Nic', 'Name', 'Phone', 'User_univ', 'User_age', 'Sex', \
        'Type', 'Dating_univ', 'Dating_age', 'agreement', 'Ticket')

print('라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작')
print('라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작')
print('라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작   라일락 매칭 시작')

for i in range(len(lilac_matching)):

    paid_person = lilac_matching[i]
    sex_condition = paid_person['Sex'].value
    univ_condition = paid_person['Dating_univ'].value
    age_condition = paid_person['Dating_age'].value

    pass_sex = list()
    pass_sex_univ = list()
    pass_sex_univ_age = list()

    print('\n')
    print(f"{paid_person['Name'].value} 매칭 시작 / 티켓 수: {paid_person['Ticket'].value}")

    # 성별 Pass
    for idx in range(len(lilac_dict_list)):
        candidate = lilac_dict_list[idx]
        if candidate['Sex'].value != sex_condition:
            pass_sex.append(candidate)
    '''
    print('1그룹')
    for row in pass_sex:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")
    '''
    if not pass_sex:
        print(f"{paid_person['Name'].value}님 ====> 1그룹 없음!!!")


    ## 성별 && 학교 Pass
    for idx in range(len(pass_sex)):
        candidate = pass_sex[idx]
        if univ_condition == 'myUniv' and paid_person['User_univ'].value == candidate['User_univ'].value :
            pass_sex_univ.append(candidate)

        elif univ_condition == 'dnt_m':
            if candidate['Dating_univ'].value == 'dnt_m' :
                pass_sex_univ.append(candidate)
            elif candidate['Dating_univ'].value == 'myUniv':
                if paid_person['User_univ'].value == candidate['User_univ'].value:
                    pass_sex_univ.append(candidate)
            elif candidate['Dating_univ'].value == 'Oth_Univ':
                if candidate['User_univ'].value != paid_person['User_univ'].value:
                    pass_sex_univ.append(candidate)

        elif univ_condition == 'Oth_Univ':
            if candidate['Dating_univ'].value == 'Oth_Univ' :
                if candidate['User_univ'].value != paid_person['User_univ'].value:
                    pass_sex_univ.append(candidate)
            elif candidate['Dating_univ'].value == 'dnt_m':
                if candidate['User_univ'].value != paid_person['User_univ'].value:
                    pass_sex_univ.append(candidate)


    print('2그룹')
    for row in pass_sex_univ:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")

    if not pass_sex_univ:
        print(f"{paid_person['Name'].value}님 ====> 2그룹 없음!!!")


    ### 성별 && 학교 && 나이 Pass
    for idx in range(len(pass_sex_univ)):
        candidate = pass_sex_univ[idx]

        if age_condition == 'dnt_m' :
            if candidate['Dating_age'].value == 'dnt_m' :
                pass_sex_univ_age.append(candidate)
            elif candidate['Dating_age'].value != 'dnt_m':
                if candidate['Dating_age'].value == paid_person['User_age'].value:
                    pass_sex_univ_age.append(candidate)
        elif age_condition != 'dnt_m':
            if candidate['Dating_age'].value == 'dnt_m' :
                if age_condition == candidate['User_age'].value:
                    pass_sex_univ_age.append(candidate)
            elif candidate['Dating_age'].value != 'dnt_m':
                if age_condition == candidate['User_age'].value:
                    if candidate['Dating_age'].value == paid_person['User_age'].value:
                        pass_sex_univ_age.append(candidate)

    '''
    ## 검색기
    if paid_person['Name'].value == "아무개" :
        print(f"{paid_person['Name'].value}님에 대해 성별, 대학교, 나이 조건 모두를 통과한 후보자 명단")
        for idx in range(len(pass_sex_univ_age)):
            print(pass_sex_univ_age[idx]['Name'].value, pass_sex_univ_age[idx]['Nic'].value, pass_sex_univ_age[idx]['Phone'].value,
                  pass_sex_univ_age[idx]['User_age'].value, pass_sex_univ_age[idx]['Ticket'].value)

    '''


    print('3그룹')
    for row in pass_sex_univ_age:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")

    if not pass_sex_univ_age:
        print(f"{paid_person['Name'].value}님 ====> 3그룹 없음!!!")



    # ticket 수 만큼 후보자 항목에서 랜덤 추출(중복 안되게 할 것)
    ticket = paid_person['Ticket'].value

    if len(pass_sex) == 0:
        print(f"{paid_person['Name'].value}님 매칭 절대 불가!!")
    elif len(pass_sex) != 0 and len(pass_sex)<ticket:
        result = random.sample(pass_sex, ticket)
        print(f"{paid_person['Name'].value}님 일부 티켓 환불 필요!!")




    if len(pass_sex_univ) == 0:
        result = random.sample(pass_sex, ticket)
        print(f"{paid_person['Name'].value}님 매칭 완료!!")
    elif len(pass_sex_univ) != 0 and len(pass_sex_univ) < ticket:
        result = random.sample(pass_sex_univ, len(pass_sex_univ))

        for second in result:
            for first in pass_sex:
                if second['Nic'] == first['Nic']:
                    pass_sex.remove(first)
        result.append(random.sample(pass_sex, ticket-len(pass_sex_univ)))
        print(f"{paid_person['Name'].value}님 매칭 완료!!")




    if len(pass_sex_univ_age) == 0:
        if len(pass_sex_univ)>= ticket:
            result = random.sample(pass_sex_univ, ticket)
            print(f"{paid_person['Name'].value}님 매칭 완료!!")
        elif len(pass_sex_univ) < ticket:
            if len(pass_sex_univ) == 0:
                result = random.sample(pass_sex, ticket)
                print(f"{paid_person['Name'].value}님 매칭 완료!!")
            else :
                result1 = random.sample(pass_sex_univ, len(pass_sex_univ))
                result2 = random.sample(pass_sex, ticket - len(pass_sex_univ))
                result = result1+result2
                print(f"{paid_person['Name'].value}님 매칭 완료!!")

    elif len(pass_sex_univ_age) != 0 and len(pass_sex_univ_age) < ticket :
        result1 = random.sample(pass_sex_univ_age, len(pass_sex_univ_age))

        for third in result1:
            for second in pass_sex_univ:
                if third['Nic'] == second['Nic']:
                     pass_sex_univ.remove(second)

        if len(pass_sex_univ) < ticket-len(pass_sex_univ_age) :
            for second in pass_sex_univ:
                for first in pass_sex:
                    if second['Nic'] == first['Nic']:
                        pass_sex.remove(first)
            result2 = random.sample(pass_sex, ticket-len(pass_sex_univ))
        else:
            result2 = random.sample(pass_sex_univ, ticket - len(pass_sex_univ_age))

        result = result2 + result1
    elif len(pass_sex_univ_age) >= ticket :
        result = random.sample(pass_sex_univ_age, ticket)
        print(f"{paid_person['Name'].value}님 매칭 완료!!")


    # 매칭 결과 console print
    print(f"{paid_person['Name'].value} 매칭 결과")
    for row in result:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")


    # 결과 시트 매칭 결과 저장하기

    lilacResult.cell(i+1, 1).value = paid_person['Phone'].value
    lilacResult.cell(i+1, 2).value = paid_person['Name'].value
    lilacResult.cell(i+1, 3).value = result[0]['Nic'].value
    lilacResult.cell(i+1, 4).value = result[0]['Phone'].value

    result_lenght = len(result)

    if result_lenght == 2 :
        lilacResult.cell(i+1, 5).value = result[1]['Nic'].value
        lilacResult.cell(i+1, 6).value = result[1]['Phone'].value
    elif result_lenght == 3 :
        lilacResult.cell(i+1, 5).value = result[1]['Nic'].value
        lilacResult.cell(i+1, 6).value = result[1]['Phone'].value
        lilacResult.cell(i+1, 7).value = result[2]['Nic'].value
        lilacResult.cell(i+1, 8).value = result[2]['Phone'].value
    elif result_lenght == 4 :
        lilacResult.cell(i+1, 5).value = result[1]['Nic'].value
        lilacResult.cell(i+1, 6).value = result[1]['Phone'].value
        lilacResult.cell(i+1, 7).value = result[2]['Nic'].value
        lilacResult.cell(i+1, 8).value = result[2]['Phone'].value
        lilacResult.cell(i+1, 9).value = result[3]['Nic'].value
        lilacResult.cell(i+1, 10).value = result[3]['Phone'].value
    elif result_lenght == 5 :
        lilacResult.cell(i+1, 5).value = result[1]['Nic'].value
        lilacResult.cell(i+1, 6).value = result[1]['Phone'].value
        lilacResult.cell(i+1, 7).value = result[2]['Nic'].value
        lilacResult.cell(i+1, 8).value = result[2]['Phone'].value
        lilacResult.cell(i+1, 9).value = result[3]['Nic'].value
        lilacResult.cell(i+1, 10).value = result[3]['Phone'].value
        lilacResult.cell(i+1, 11).value = result[4]['Nic'].value
        lilacResult.cell(i+1, 12).value = result[4]['Phone'].value


    wb.save('result40000.xlsx')




print('데이지 이성 매칭 시작   데이지 이성 매칭 시작   데이지 이성 매칭 시작   데이지 이성 매칭 시작')
print('데이지 이성 매칭 시작   데이지 이성 매칭 시작   데이지 이성 매칭 시작   데이지 이성 매칭 시작')
print('데이지 이성 매칭 시작   데이지 이성 매칭 시작   데이지 이성 매칭 시작   데이지 이성 매칭 시작')

for i in range(len(daisy_matching)):
    paid_person = daisy_matching[i]
    sex_condition = paid_person['Sex'].value
    univ_condition = paid_person['Dating_univ'].value
    age_condition = paid_person['Dating_age'].value

    pass_sex = list()
    pass_sex_univ = list()
    pass_sex_univ_age = list()

    print('\n')
    print(f"{paid_person['Name'].value} 매칭 시작 / 티켓 수: {paid_person['Ticket'].value}")

    # 성별 PASS ==> 1그룹
    for idx in range(len(daisy_dict_list)):
        candidate = daisy_dict_list[idx]
        if candidate['Sex'].value != sex_condition:
            pass_sex.append(candidate)
    '''
    print('1그룹')
    for row in pass_sex:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")
    '''
    if not pass_sex:
        print(f"{paid_person['Name'].value}님 ====> 1그룹 없음!!!")


    ## 성별 && 학교 Pass ==> 2그룹
    for idx in range(len(pass_sex)):
        candidate = pass_sex[idx]
        if univ_condition == 'myUniv' and paid_person['User_univ'].value == candidate['User_univ'].value :
            pass_sex_univ.append(candidate)

        elif univ_condition == 'dnt_m':
            if candidate['Dating_univ'].value == 'dnt_m' :
                pass_sex_univ.append(candidate)
            elif candidate['Dating_univ'].value == 'myUniv':
                if paid_person['User_univ'].value == candidate['User_univ'].value:
                    pass_sex_univ.append(candidate)
            elif candidate['Dating_univ'].value == 'Oth_Univ':
                if candidate['User_univ'].value != paid_person['User_univ'].value:
                    pass_sex_univ.append(candidate)

        elif univ_condition == 'Oth_Univ':
            if candidate['Dating_univ'].value == 'Oth_Univ' :
                if candidate['User_univ'].value != paid_person['User_univ'].value:
                    pass_sex_univ.append(candidate)
            elif candidate['Dating_univ'].value == 'dnt_m':
                if candidate['User_univ'].value != paid_person['User_univ'].value:
                    pass_sex_univ.append(candidate)

    print('2그룹')
    for row in pass_sex_univ:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")

    if not pass_sex_univ:
        print(f"{paid_person['Name'].value}님 ====> 2그룹 없음!!!")


    ### 성별 && 학교 && 나이 Pass ==> 3그룹
    for idx in range(len(pass_sex_univ)):
        candidate = pass_sex_univ[idx]

        if age_condition == 'dnt_m' :
            if candidate['Dating_age'].value == 'dnt_m' :
                pass_sex_univ_age.append(candidate)
            elif candidate['Dating_age'].value != 'dnt_m':
                if candidate['Dating_age'].value == paid_person['User_age'].value:
                    pass_sex_univ_age.append(candidate)

        elif age_condition != 'dnt_m':

            if candidate['Dating_age'].value == 'dnt_m' :
                if age_condition == candidate['User_age'].value:
                    pass_sex_univ_age.append(candidate)

            elif candidate['Dating_age'].value != 'dnt_m':
                if age_condition == candidate['User_age'].value:
                    if candidate['Dating_age'].value == paid_person['User_age'].value:
                        pass_sex_univ_age.append(candidate)


    print('3그룹')
    for row in pass_sex_univ_age:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : {row['Phone'].value}")

    if not pass_sex_univ_age:
        print(f"{paid_person['Name'].value}님 ====> 3그룹 없음!!!")



    # ticket 수 만큼 후보자 항목에서 랜덤 추출(중복 안되게 할 것)
    ticket = paid_person['Ticket'].value

    if len(pass_sex) == 0:
        print(f"{paid_person['Name'].value}님 매칭 절대 불가!!")
    elif len(pass_sex) != 0 and len(pass_sex)<ticket:
        result = random.sample(pass_sex, ticket)
        print(f"{paid_person['Name'].value}님 일부 티켓 환불 필요!!")




    if len(pass_sex_univ) == 0:
        result = random.sample(pass_sex, ticket)
        print(f"{paid_person['Name'].value}님 매칭 완료!!")
    elif len(pass_sex_univ) != 0 and len(pass_sex_univ) < ticket:
        result = random.sample(pass_sex_univ, len(pass_sex_univ))

        for second in result:
            for first in pass_sex:
                if second['Nic'] == first['Nic']:
                    pass_sex.remove(first)
        result.append(random.sample(pass_sex, ticket-len(pass_sex_univ)))
        print(f"{paid_person['Name'].value}님 매칭 완료!!")




    if len(pass_sex_univ_age) == 0:
        if len(pass_sex_univ)>= ticket:
            result = random.sample(pass_sex_univ, ticket)
            print(f"{paid_person['Name'].value}님 매칭 완료!!")
        elif len(pass_sex_univ) < ticket:
            if len(pass_sex_univ) == 0:
                result = random.sample(pass_sex, ticket)
                print(f"{paid_person['Name'].value}님 매칭 완료!!")
            else :
                result1 = random.sample(pass_sex_univ, len(pass_sex_univ))
                result2 = random.sample(pass_sex, ticket - len(pass_sex_univ))
                result = result1+result2
                print(f"{paid_person['Name'].value}님 매칭 완료!!")

    elif len(pass_sex_univ_age) != 0 and len(pass_sex_univ_age) < ticket :
        result1 = random.sample(pass_sex_univ_age, len(pass_sex_univ_age))

        for third in result1:
            for second in pass_sex_univ:
                if third['Nic'] == second['Nic']:
                     pass_sex_univ.remove(second)

        if len(pass_sex_univ) < ticket-len(pass_sex_univ_age) :
            for second in pass_sex_univ:
                for first in pass_sex:
                    if second['Nic'] == first['Nic']:
                        pass_sex.remove(first)
            result2 = random.sample(pass_sex, ticket-len(pass_sex_univ))
        else:
            result2 = random.sample(pass_sex_univ, ticket - len(pass_sex_univ_age))

        result = result2 + result1
    elif len(pass_sex_univ_age) >= ticket :
        result = random.sample(pass_sex_univ_age, ticket)
        print(f"{paid_person['Name'].value}님 매칭 완료!!")


    # 매칭 결과 console print
    print(f"{paid_person['Name'].value} 매칭 결과")
    for row in result:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : 0{row['Phone'].value}")



    # 결과 시트 매칭 결과 저장하기

    daisyResult.cell(i + 1, 1).value = paid_person['Phone'].value
    daisyResult.cell(i + 1, 2).value = paid_person['Name'].value
    daisyResult.cell(i + 1, 3).value = result[0]['Nic'].value
    daisyResult.cell(i + 1, 4).value = result[0]['Phone'].value

    result_lenght = len(result)

    if result_lenght == 2:
        daisyResult.cell(i + 1, 5).value = result[1]['Nic'].value
        daisyResult.cell(i + 1, 6).value = result[1]['Phone'].value
    elif result_lenght == 3:
        daisyResult.cell(i + 1, 5).value = result[1]['Nic'].value
        daisyResult.cell(i + 1, 6).value = result[1]['Phone'].value
        daisyResult.cell(i + 1, 7).value = result[2]['Nic'].value
        daisyResult.cell(i + 1, 8).value = result[2]['Phone'].value
    elif result_lenght == 4:
        daisyResult.cell(i + 1, 5).value = result[1]['Nic'].value
        daisyResult.cell(i + 1, 6).value = result[1]['Phone'].value
        daisyResult.cell(i + 1, 7).value = result[2]['Nic'].value
        daisyResult.cell(i + 1, 8).value = result[2]['Phone'].value
        daisyResult.cell(i + 1, 9).value = result[3]['Nic'].value
        daisyResult.cell(i + 1, 10).value = result[3]['Phone'].value
    elif result_lenght == 5:
        daisyResult.cell(i + 1, 5).value = result[1]['Nic'].value
        daisyResult.cell(i + 1, 6).value = result[1]['Phone'].value
        daisyResult.cell(i + 1, 7).value = result[2]['Nic'].value
        daisyResult.cell(i + 1, 8).value = result[2]['Phone'].value
        daisyResult.cell(i + 1, 9).value = result[3]['Nic'].value
        daisyResult.cell(i + 1, 10).value = result[3]['Phone'].value
        daisyResult.cell(i + 1, 11).value = result[4]['Nic'].value
        daisyResult.cell(i + 1, 12).value = result[4]['Phone'].value

    wb.save('result40000.xlsx')

sys.stdout.close()


print('데이지 동성 매칭 시작   데이지 동성 매칭 시작   데이지 동성 매칭 시작   데이지 동성 매칭 시작')
print('데이지 동성 매칭 시작   데이지 동성 매칭 시작   데이지 동성 매칭 시작   데이지 동성 매칭 시작')
print('데이지 동성 매칭 시작   데이지 동성 매칭 시작   데이지 동성 매칭 시작   데이지 동성 매칭 시작')

for i in range(len(clover_matching)):

    paid_person = clover_matching[i]
    sex_condition = paid_person['Sex'].value
    univ_condition = paid_person['Dating_univ'].value
    age_condition = paid_person['Dating_age'].value
    phone_num = paid_person['Phone'].value

    pass_sex = list()
    pass_sex_univ = list()
    pass_sex_univ_age = list()

    print('\n')
    print(f"{paid_person['Name'].value} 매칭 시작 / 티켓 수: {paid_person['Ticket'].value}")

    # 성별 Pass
    for idx in range(len(clover_dict_list)):
        candidate = clover_dict_list[idx]
        if candidate['Sex'].value == sex_condition:
            if candidate['Phone'].value != phone_num:
                pass_sex.append(candidate)
    '''
    print('1그룹')
    for row in pass_sex:
        print(row)
    '''
    if not pass_sex:
        print(f"{paid_person['Name'].value}님 ====> 1그룹 없음!!!")


    ## 성별 && 학교 Pass
    for idx in range(len(pass_sex)):
        candidate = pass_sex[idx]
        if univ_condition == 'myUniv' and paid_person['User_univ'].value == candidate['User_univ'].value:
            pass_sex_univ.append(candidate)
        elif univ_condition == 'dnt_m' and candidate['Dating_univ'].value == 'dnt_m':
            pass_sex_univ.append(candidate)
        elif univ_condition == 'dnt_m' and paid_person['User_univ'].value == candidate['Dating_univ'].value:
            pass_sex_univ.append(candidate)
        elif univ_condition == 'elseUniv' and paid_person['User_univ'].value != candidate['User_univ'].value :
            pass_sex_univ.append(candidate)
    '''
    print('2그룹')
    for row in pass_sex_univ:
        print(row)
    '''
    if not pass_sex_univ:
        print(f"{paid_person['Name'].value}님 ====> 2그룹 없음!!!")


    ### 성별 && 학교 && 나이 Pass
    for idx in range(len(pass_sex_univ)):
        candidate = pass_sex_univ[idx]
        if age_condition == 'dnt_m' and candidate['Dating_age'].value == 'dnt_m':
            pass_sex_univ_age.append(candidate)
        elif age_condition == 'dnt_m' and candidate['Dating_age'].value == paid_person['User_age'].value:
            pass_sex_univ_age.append(candidate)
        elif age_condition != 'dnt_m' and age_condition == candidate['User_age'].value \
                and candidate['Dating_age'].value == paid_person['User_age'].value:
            pass_sex_univ_age.append(candidate)


    '''
    print('3그룹')
    for row in pass_sex_univ_age:
        print(row)
    '''
    if not pass_sex_univ_age:
        print(f"{paid_person['Name'].value}님 ====> 3그룹 없음!!!")



    # ticket 수 만큼 후보자 항목에서 랜덤 추출(중복 안되게 할 것)

    ticket = paid_person['Ticket'].value

    if len(pass_sex) == 0:
        print(f"{paid_person['Name'].value}님 매칭 절대 불가!!")
    elif len(pass_sex) != 0 and len(pass_sex)<ticket:
        result = random.sample(pass_sex, ticket)
        print(f"{paid_person['Name'].value}님 일부 티켓 환불 필요!!")




    if len(pass_sex_univ) == 0:
        result = random.sample(pass_sex, ticket)
        print(f"{paid_person['Name'].value}님 매칭 완료!!")
    elif len(pass_sex_univ) != 0 and len(pass_sex_univ) < ticket:
        result = random.sample(pass_sex_univ, len(pass_sex_univ))

        for second in result:
            for first in pass_sex:
                if second['Nic'] == first['Nic']:
                    pass_sex.remove(first)
        result.append(random.sample(pass_sex, ticket-len(pass_sex_univ)))
        print(f"{paid_person['Name'].value}님 매칭 완료!!")



    if len(pass_sex_univ_age) == 0:
        if len(pass_sex_univ)>= ticket:
            result = random.sample(pass_sex_univ, ticket)
            print(f"{paid_person['Name'].value}님 매칭 완료!!")
        elif len(pass_sex_univ) < ticket:
            if len(pass_sex_univ) == 0:
                result = random.sample(pass_sex, ticket)
                print(f"{paid_person['Name'].value}님 매칭 완료!!")
            else :
                result1 = random.sample(pass_sex_univ, len(pass_sex_univ))
                result2 = random.sample(pass_sex, ticket - len(pass_sex_univ))
                result = result1+result2
                print(f"{paid_person['Name'].value}님 매칭 완료!!")


    elif len(pass_sex_univ_age) != 0 and len(pass_sex_univ_age) < ticket :
        result1 = random.sample(pass_sex_univ_age, len(pass_sex_univ_age))

        for third in result1:
            for second in pass_sex_univ:
                if third['Nic'] == second['Nic']:
                     pass_sex_univ.remove(second)

        if len(pass_sex_univ) < ticket-len(pass_sex_univ_age) :
            for second in pass_sex_univ:
                for first in pass_sex:
                    if second['Nic'] == first['Nic']:
                        pass_sex.remove(first)
            result2 = random.sample(pass_sex, ticket-len(pass_sex_univ))
        else:
            result2 = random.sample(pass_sex_univ, ticket - len(pass_sex_univ_age))

        result = result2 + result1
    elif len(pass_sex_univ_age) >= ticket :
        result = random.sample(pass_sex_univ_age, ticket)
        print(f"{paid_person['Name'].value}님 매칭 완료!!")


    # 매칭 결과 console print
    print(f"{paid_person['Name'].value} 매칭 결과")
    for row in result:
        print(f"닉네임 : {row['Nic'].value} // 전화번호 : 0{row['Phone'].value}")



    # 결과 시트 매칭 결과 저장하기

    cloverResult.cell(i + 1, 1).value = paid_person['Phone'].value
    cloverResult.cell(i + 1, 2).value = paid_person['Name'].value
    cloverResult.cell(i + 1, 3).value = result[0]['Nic'].value
    cloverResult.cell(i + 1, 4).value = result[0]['Phone'].value

    result_lenght = len(result)

    if result_lenght == 2:
        cloverResult.cell(i + 1, 5).value = result[1]['Nic'].value
        cloverResult.cell(i + 1, 6).value = result[1]['Phone'].value
    elif result_lenght == 3:
        cloverResult.cell(i + 1, 5).value = result[1]['Nic'].value
        cloverResult.cell(i + 1, 6).value = result[1]['Phone'].value
        cloverResult.cell(i + 1, 7).value = result[2]['Nic'].value
        cloverResult.cell(i + 1, 8).value = result[2]['Phone'].value
    elif result_lenght == 4:
        cloverResult.cell(i + 1, 5).value = result[1]['Nic'].value
        cloverResult.cell(i + 1, 6).value = result[1]['Phone'].value
        cloverResult.cell(i + 1, 7).value = result[2]['Nic'].value
        cloverResult.cell(i + 1, 8).value = result[2]['Phone'].value
        cloverResult.cell(i + 1, 9).value = result[3]['Nic'].value
        cloverResult.cell(i + 1, 10).value = result[3]['Phone'].value
    elif result_lenght == 5:
        cloverResult.cell(i + 1, 5).value = result[1]['Nic'].value
        cloverResult.cell(i + 1, 6).value = result[1]['Phone'].value
        cloverResult.cell(i + 1, 7).value = result[2]['Nic'].value
        cloverResult.cell(i + 1, 8).value = result[2]['Phone'].value
        cloverResult.cell(i + 1, 9).value = result[3]['Nic'].value
        cloverResult.cell(i + 1, 10).value = result[3]['Phone'].value
        cloverResult.cell(i + 1, 11).value = result[4]['Nic'].value
        cloverResult.cell(i + 1, 12).value = result[4]['Phone'].value

    wb.save('result40000.xlsx')

print('@@@@@ 플로스팅 4회차 종료 @@@@@')
print('@@@@@ 플로스팅 4회차 종료 @@@@@')
print('@@@@@ 플로스팅 4회차 종료 @@@@@')
print('@@@@@ 플로스팅 4회차 종료 @@@@@')
print('@@@@@ 플로스팅 4회차 종료 @@@@@')